"""用 openpyxl 生成发货单 xlsx，复用已验证可打印的 Excel 打印配置。

关键参数来自《打印表(1).xlsx》实测：
  paperSize=132  → EPSON 连续纸专用代码
  orientation=landscape
  scale=110
  margins: left=0.078in(2mm), right=1.102in(28mm), top/bottom=1.0in(25.4mm)
"""

import os
import tempfile

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_00


# ── 样式常量 ────────────────────────────────────────────────────────────────

def _font(size=11, bold=False, name="宋体"):
    return Font(name=name, size=size, bold=bold)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border_bottom(thick=False):
    side = Side(style="medium" if thick else "thin")
    return Border(bottom=side)


def _full_border(top=False, bottom=False, thick_bottom=False):
    thin = Side(style="thin")
    med  = Side(style="medium")
    return Border(
        top    = med if top  else None,
        bottom = med if thick_bottom else (thin if bottom else None),
        left   = None,
        right  = None,
    )


# ── 主函数 ──────────────────────────────────────────────────────────────────

def render_xlsx(data: dict, out_path: str) -> str:
    """把订单数据写入 xlsx，配置 EPSON 连续纸打印参数。

    data 包含：brand_name, customer, order_id, created_at,
               items (list of dict), total
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发货单"

    # ── 打印页面设置（完全复制参考文件的工作配置） ──
    ws.page_setup.orientation   = "landscape"
    ws.page_setup.paperSize     = 132       # EPSON 连续纸代码，WPS/Excel 均可识别
    ws.page_setup.scale         = 110       # 110% 缩放，与参考文件一致
    ws.page_setup.horizontalDpi = 600

    # 页边距（单位：英寸），与参考文件完全一致
    ws.page_margins.left   = 0.0784722222  # ≈2mm
    ws.page_margins.right  = 1.10208333    # ≈28mm（针孔边缘留白）
    ws.page_margins.top    = 1.0           # ≈25.4mm
    ws.page_margins.bottom = 1.0           # ≈25.4mm
    ws.page_margins.header = 0.5
    ws.page_margins.footer = 0.5

    # ── 列宽（单位：字符宽度），按参考文件微调 ──
    ws.column_dimensions["A"].width =  5.4   # #
    ws.column_dimensions["B"].width = 22.0   # 品名（加宽）
    ws.column_dimensions["C"].width =  8.0   # 规格
    ws.column_dimensions["D"].width =  7.0   # 数量
    ws.column_dimensions["E"].width =  9.0   # 单价
    ws.column_dimensions["F"].width = 11.0   # 小计

    # ── 第1行：标题 ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22.5
    ws["A1"] = f"{data['brand_name']}  发货单"
    ws["A1"].font      = _font(size=16, bold=True)
    ws["A1"].alignment = _align("left")

    # ── 第2行：客户信息 ───────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 14.0
    ws["A2"] = (
        f"客户：{data['customer']}    "
        f"单号：#{data['order_id']}    "
        f"日期：{data['created_at']}"
    )
    ws["A2"].font      = _font(size=11)
    ws["A2"].alignment = _align("left")

    # ── 第3行：表头 ───────────────────────────────────────────────────────────
    headers = ["#", "品名", "规格", "数量", "单价", "小计"]
    aligns  = ["center", "left", "center", "right", "right", "right"]
    ws.row_dimensions[3].height = 15.0
    for col_idx, (h, al) in enumerate(zip(headers, aligns), start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font      = _font(size=11, bold=True)
        cell.alignment = _align(al)
        cell.border    = _full_border(top=True, bottom=True)

    # ── 数据行 ────────────────────────────────────────────────────────────────
    row = 4
    items = data.get("items", [])
    for item in items:
        ws.row_dimensions[row].height = 13.0
        is_rep = item.get("is_replacement", False)

        def _set(col, val, h="left", bold=False, num_fmt=None):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = _font(size=11, bold=bold)
            c.alignment = _align(h)
            c.border    = _border_bottom()
            if num_fmt:
                c.number_format = num_fmt
            return c

        _set(1, item.get("index", row - 3), h="center")
        name_val = item.get("product_name", "")
        if is_rep:
            name_val += "  [补]"
        _set(2, name_val, h="left")
        _set(3, item.get("spec") or "", h="center")
        _set(4, item.get("qty", 0), h="right")
        _set(5, 0 if is_rep else item.get("price", 0),    h="right", num_fmt="0.00")
        _set(6, 0 if is_rep else item.get("subtotal", 0), h="right", num_fmt="0.00")
        row += 1

    # 空行补足（保证合计行位置相对稳定）
    while row < 20:
        ws.row_dimensions[row].height = 13.0
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = _border_bottom()
        row += 1

    # ── 合计行 ────────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 15.0
    total_label = ws.cell(row=row, column=5, value="合计")
    total_label.font      = _font(size=11, bold=True)
    total_label.alignment = _align("right")
    total_label.border    = _full_border(top=True)

    total_val = ws.cell(row=row, column=6, value=data.get("total", 0))
    total_val.font         = _font(size=12, bold=True)
    total_val.alignment    = _align("right")
    total_val.number_format = "0.00"
    total_val.border       = _full_border(top=True)

    # ── 签收行 ────────────────────────────────────────────────────────────────
    row += 1
    ws.row_dimensions[row].height = 15.0
    sig = ws.cell(row=row, column=1, value="制单：____________    签收：____________")
    sig.font      = _font(size=11)
    sig.alignment = _align("left")

    wb.save(out_path)
    return out_path
